import uuid
from io import BytesIO
from datetime import datetime

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.ext.asyncio import AsyncSession

from Backend.API_Layer.interfaces.permenent_employee_details_interfaces import (
    CreatePermanentEmployeeRequest,
    CreatePermanentEmployeeResponse,
    UpdatePermanentEmployeeRequest,
)

from Backend.Business_Layer.utils.excel_parcer import parse_excel
from Backend.Business_Layer.utils.uuid_generator import generate_uuid7
from Backend.DAL.dao.permanent_employee_details_dao import PermanentEmployeeDetailsDAO
from Backend.DAL.models.models import EmployeeDetails


class PermanentEmployeeDetailsService:

    CURRENCIES = ["INR", "USD", "EUR", "GBP", "AED", "SGD", "AUD", "CAD"]

    EMPLOYEE_TYPES = ["Full-Time", "Part-Time", "Intern", "Contractor", "Freelance"]

    WORK_MODES = ["Office", "Remote", "Hybrid"]

    EMPLOYMENT_STATUSES = [
        "Probation",
        "Active",
        "Resigned",
        "Terminated",
        "Absconded",
        "Exited",
        "On-Notice",
    ]

    BLOOD_GROUPS = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]

    GENDERS = ["Male", "Female", "Other"]

    MARITAL_STATUSES = ["Single", "Married", "Divorced", "Widowed"]

    def __init__(self):
        self.dao = PermanentEmployeeDetailsDAO()

    # =========================================================
    # CREATE EMPLOYEE HELPERS
    # =========================================================

    async def generate_employee_id(self, db: AsyncSession):

        last_employee_id = await self.dao.get_last_employee(db)

        if last_employee_id:
            new_employee_id = last_employee_id + 1
        else:
            new_employee_id = 5100001

        return str(new_employee_id)

    async def generate_work_email(
        self, db: AsyncSession, first_name: str, last_name: str
    ):

        domain = "pavestechnologies.com"

        first_parts = first_name.lower().split()

        last = last_name.lower()

        combinations = []

        combinations.append("".join(first_parts) + "." + last)

        combinations.append(first_parts[0] + "." + last)

        if len(first_parts) > 1:
            combinations.append(first_parts[0] + first_parts[1][0] + "." + last)

        combinations.append(first_parts[0][0] + last)

        for combo in combinations:

            email = f"{combo}@{domain}"

            existing = await self.dao.get_employee_by_email(db, email)

            if not existing:
                return email

        base = combinations[0]

        counter = 1

        while True:

            email = f"{base}{counter}@{domain}"

            existing = await self.dao.get_employee_by_email(db, email)

            if not existing:
                return email

            counter += 1

    async def resolve_reporting_manager_employee_id(
        self, db: AsyncSession, reporting_manager
    ):

        if reporting_manager is None:
            return None

        if str(reporting_manager).strip() == "":
            return None

        manager = await self.dao.get_employee_by_manager_value(db, reporting_manager)

        if not manager:
            raise ValueError("Invalid reporting manager selected")

        return manager.employee_id

    # =========================================================
    # CREATE EMPLOYEE
    # =========================================================

    async def create_employee(
        self,
        db: AsyncSession,
        payload: CreatePermanentEmployeeRequest,
        current_user_id: str,
    ):

        existing = await self.dao.get_employee_by_user_uuid(db, payload.user_uuid)

        if existing:
            raise ValueError("Employee already created for this user")

        employee_id = await self.generate_employee_id(db)

        work_email = await self.generate_work_email(
            db, payload.first_name, payload.last_name
        )

        reporting_manager_employee_id = (
            await self.resolve_reporting_manager_employee_id(
                db, payload.reporting_manager_uuid
            )
        )

        employee = EmployeeDetails(
            employee_uuid=str(uuid.uuid4()),
            user_uuid=payload.user_uuid,
            employee_id=employee_id,
            first_name=payload.first_name,
            middle_name=payload.middle_name,
            last_name=payload.last_name,
            date_of_birth=payload.date_of_birth,
            work_email=work_email,
            contact_number=payload.contact_number,
            department_uuid=payload.department_uuid,
            designation_uuid=payload.designation_uuid,
            reporting_manager_uuid=(reporting_manager_employee_id),
            employment_type=payload.employment_type,
            joining_date=payload.joining_date,
            location=payload.location,
            work_mode=payload.work_mode,
            employment_status=payload.employment_status,
            blood_group=payload.blood_group,
            gender=payload.gender,
            marital_status=payload.marital_status,
            total_experience=payload.total_experience,
            created_by=current_user_id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )

        employee = await self.dao.create_employee(db, employee)

        await db.commit()

        return CreatePermanentEmployeeResponse(
            employee_uuid=employee.employee_uuid,
            employee_id=employee.employee_id,
            work_email=employee.work_email,
            message="Employee created successfully",
        )

    # =========================================================
    # GET EMPLOYEE
    # =========================================================

    async def get_employee_by_uuid(self, db: AsyncSession, employee_uuid: str):

        employee = await self.dao.get_employee_by_uuid(db, employee_uuid)

        if not employee:
            raise ValueError("Employee not found")

        return {
            "user_uuid": employee.user_uuid,
            "employee_uuid": employee.employee_uuid,
            "employee_id": employee.employee_id,
            "first_name": employee.first_name,
            "middle_name": employee.middle_name,
            "last_name": employee.last_name,
            "date_of_birth": employee.date_of_birth,
            "work_email": employee.work_email,
            "contact_number": employee.contact_number,
            "department_uuid": employee.department_uuid,
            "designation_uuid": employee.designation_uuid,
            "reporting_manager_uuid": employee.reporting_manager_uuid,
            "employment_type": employee.employment_type,
            "joining_date": employee.joining_date,
            "location": employee.location,
            "work_mode": employee.work_mode,
            "employment_status": employee.employment_status,
            "blood_group": employee.blood_group,
            "gender": employee.gender,
            "marital_status": employee.marital_status,
            "total_experience": employee.total_experience,
        }

    # =========================================================
    # GET ALL EMPLOYEES
    # =========================================================

    async def get_all_employees(self, db: AsyncSession):

        employees = await self.dao.get_all_employees(db)

        response = []

        for emp in employees:

            response.append(
                {
                    "user_uuid": emp["user_uuid"],
                    "employee_uuid": emp["employee_uuid"],
                    "employee_id": emp.get("employee_id"),
                    "first_name": emp.get("first_name"),
                    "middle_name": emp.get("middle_name"),
                    "last_name": emp.get("last_name"),
                    "date_of_birth": (
                        str(emp["date_of_birth"]) if emp.get("date_of_birth") else None
                    ),
                    "work_email": emp.get("work_email"),
                    "contact_number": emp.get("contact_number"),
                    "department_uuid": emp.get("department_uuid"),
                    "designation_uuid": emp.get("designation_uuid"),
                    "reporting_manager_uuid": (emp.get("reporting_manager_uuid")),
                    "employment_type": emp.get("employment_type"),
                    "joining_date": (
                        str(emp["joining_date"]) if emp.get("joining_date") else None
                    ),
                    "location": emp.get("location"),
                    "work_mode": emp.get("work_mode"),
                    "employment_status": (emp.get("employment_status")),
                    "blood_group": emp.get("blood_group"),
                    "gender": emp.get("gender"),
                    "marital_status": emp.get("marital_status"),
                    "total_experience": emp.get("total_experience"),
                    "bg_status": emp.get("bg_status", "NOT_STARTED"),
                    "bgv_status": emp.get("bgv_status"),
                }
            )

        return response

    # =========================================================
    # UPDATE EMPLOYEE
    # =========================================================

    async def update_employee(
        self,
        db: AsyncSession,
        employee_uuid: str,
        request: UpdatePermanentEmployeeRequest,
    ):

        employee = await self.dao.get_employee_by_uuid(db, employee_uuid)

        if not employee:
            raise ValueError("Employee not found")

        if request.first_name is not None:
            employee.first_name = request.first_name

        if request.middle_name is not None:
            employee.middle_name = request.middle_name

        if request.last_name is not None:
            employee.last_name = request.last_name

        if request.date_of_birth is not None:
            employee.date_of_birth = request.date_of_birth

        if request.contact_number is not None:
            employee.contact_number = request.contact_number

        if request.department_uuid is not None:
            employee.department_uuid = request.department_uuid

        if request.designation_uuid is not None:
            employee.designation_uuid = request.designation_uuid

        if request.reporting_manager_uuid is not None:

            employee.reporting_manager_uuid = (
                await self.resolve_reporting_manager_employee_id(
                    db, request.reporting_manager_uuid
                )
            )

        if request.employment_type is not None:
            employee.employment_type = request.employment_type

        if request.joining_date is not None:
            employee.joining_date = request.joining_date

        if request.location is not None:
            employee.location = request.location

        if request.work_mode is not None:
            employee.work_mode = request.work_mode

        if request.employment_status is not None:
            employee.employment_status = request.employment_status

        if request.blood_group is not None:
            employee.blood_group = request.blood_group

        if request.gender is not None:
            employee.gender = request.gender

        if request.marital_status is not None:
            employee.marital_status = request.marital_status

        if request.total_experience is not None:
            employee.total_experience = request.total_experience

        employee.updated_at = datetime.utcnow()

        employee = await self.dao.update_employee(db, employee)

        await db.commit()

        return {
            "employee_uuid": employee.employee_uuid,
            "employee_id": employee.employee_id,
            "message": "Employee updated successfully",
        }

    # =========================================================
    # UPDATE EMPLOYEE BGV STATUS
    # =========================================================

    async def update_bgv_status(self, db: AsyncSession, user_uuid: str, bgv_status: str):

        employee = await self.dao.get_employee_by_user_uuid(db, user_uuid)

        if not employee:
            raise ValueError("Employee not found")

        await self.dao.update_bgv_status(db, user_uuid, bgv_status)

        await db.commit()

        return {
            "user_uuid": user_uuid,
            "bgv_status": bgv_status,
            "message": "BGV status updated successfully",
        }

    # =========================================================
    # DELETE EMPLOYEE
    # =========================================================

    async def delete_employee(self, db: AsyncSession, employee_uuid: str):

        employee = await self.dao.get_employee_by_uuid(db, employee_uuid)

        if not employee:
            raise ValueError("Employee not found")

        await self.dao.delete_employee(db, employee_uuid)

        await db.commit()

        return {"message": "Employee deleted successfully"}

    # =========================================================
    # BULK HELPER
    # =========================================================

    def get_employee_id(self, manager_value):

        if not manager_value:
            return None

        value = str(manager_value).strip()

        if " - " in value:
            return value.split(" - ", 1)[0].strip()

        return value

    # =========================================================
    # BULK DIRECT UPLOAD
    # =========================================================

    async def bulk_direct_upload(self, db, file, current_user_id):

        data = parse_excel(file)

        success_count = 0

        failed_records = []

        uploaded_employee_ids = set()

        uploaded_emails = set()

        for index, row in enumerate(data):

            try:

                # ============================================
                # DATE CONVERSION
                # ============================================

                for field in ["joining_date", "date_of_birth"]:

                    if row.get(field):

                        value = row.get(field)

                        if isinstance(value, str):

                            try:
                                row[field] = datetime.strptime(
                                    value.strip(), "%d/%m/%Y"
                                ).date()

                            except ValueError:
                                pass

                employee_id = row.get("employee_id")

                work_email = row.get("work_email")

                # ============================================
                # EXCEL DUPLICATE VALIDATION
                # ============================================

                if employee_id in uploaded_employee_ids:
                    raise Exception(
                        f"Duplicate employee_id in excel: " f"{employee_id}"
                    )

                if work_email in uploaded_emails:
                    raise Exception(f"Duplicate work_email in excel: " f"{work_email}")

                uploaded_employee_ids.add(employee_id)

                uploaded_emails.add(work_email)

                # ============================================
                # DATABASE DUPLICATE VALIDATION
                # ============================================

                employee_exists = await self.dao.check_employee_id_exists(
                    db, employee_id
                )

                if employee_exists:
                    raise Exception(f"Employee ID already exists: " f"{employee_id}")

                email_exists = await self.dao.check_work_email_exists(db, work_email)

                if email_exists:
                    raise Exception(f"Work email already exists: " f"{work_email}")

                # ============================================
                # DEPARTMENT
                # ============================================

                department_uuid = await self.dao.get_department_uuid(
                    db, row.get("department")
                )

                if not department_uuid:
                    raise Exception(f"Invalid department: " f"{row.get('department')}")

                # ============================================
                # DESIGNATION
                # ============================================

                designation_uuid = await self.dao.get_designation_uuid(
                    db, row.get("designation")
                )

                if not designation_uuid:
                    raise Exception(
                        f"Invalid designation: " f"{row.get('designation')}"
                    )

                # ============================================
                # MANAGER
                # ============================================

                reporting_manager_employee_id = self.get_employee_id(
                    row.get("reporting_manager_uuid")
                )

                if reporting_manager_employee_id:

                    manager_exists = await self.dao.check_employee_id_exists(
                        db, reporting_manager_employee_id
                    )

                    if not manager_exists:
                        raise Exception(
                            f"Invalid reporting manager: "
                            f"{reporting_manager_employee_id}"
                        )

                # ============================================
                # UUIDS
                # ============================================

                user_uuid = str(generate_uuid7())

                employee_uuid = str(generate_uuid7())

                personal_uuid = str(generate_uuid7())

                bank_uuid = str(generate_uuid7())

                pf_uuid = str(generate_uuid7())

                current_address_uuid = str(generate_uuid7())

                permanent_address_uuid = str(generate_uuid7())

                # ============================================
                # COUNTRY UUID
                # ============================================

                country_uuid = None

                if row.get("country_code"):

                    country_uuid = await self.dao.get_country_uuid_by_calling_code(
                        db, row.get("country_code")
                    )

                # ============================================
                # TRANSACTION
                # ============================================

                async with db.begin_nested():

                    await self.dao.insert_offer_letter(
                        db,
                        row,
                        user_uuid,
                        current_user_id,
                        reporting_manager_employee_id,
                    )

                    await self.dao.insert_employee(
                        db,
                        row,
                        user_uuid,
                        employee_uuid,
                        employee_id,
                        work_email,
                        department_uuid,
                        designation_uuid,
                        reporting_manager_employee_id,
                        current_user_id,
                    )

                    await self.dao.insert_personal_details(
                        db, personal_uuid, user_uuid, row
                    )

                    await self.dao.insert_bank_details(db, bank_uuid, user_uuid)

                    await self.dao.insert_pf_details(db, pf_uuid, user_uuid)

                    await self.dao.insert_address(
                        db, current_address_uuid, user_uuid, "current", country_uuid
                    )

                    await self.dao.insert_address(
                        db, permanent_address_uuid, user_uuid, "permanent", country_uuid
                    )

                    await db.flush()

                success_count += 1

            except Exception as e:

                failed_records.append(
                    {
                        "row": index + 2,
                        "reason": str(e),
                    }
                )

        await db.commit()

        return {
            "message": "Bulk upload completed",
            "success_count": success_count,
            "failed_count": len(failed_records),
            "failed_records": failed_records,
        }

    # =========================================================
    # DOWNLOAD TEMPLATE
    # =========================================================

    async def download_bulk_template(self, db):

        departments = await self.dao.get_departments(db)

        designations = await self.dao.get_designations(db)

        country_codes = await self.dao.get_country_codes(db)

        employees = await self.dao.get_employees_for_dropdown(db)

        manager_values = []

        for emp in employees:

            employee_id, first_name, middle_name, last_name = emp

            name = " ".join(
                part for part in [first_name, middle_name, last_name] if part
            ).strip()

            manager_values.append(f"{employee_id} - {name}")

        headers = [
            "first_name",
            "middle_name",
            "last_name",
            "mail",
            "country_code",
            "contact_number",
            "employee_type",
            "currency",
            "joining_date",
            "cc_emails",
            "total_ctc",
            "job_id",
            "date_of_birth",
            "work_email",
            "employee_id",
            "department",
            "designation",
            "reporting_manager_uuid",
            "employment_type",
            "location",
            "work_mode",
            "employment_status",
            "blood_group",
            "gender",
            "marital_status",
            "total_experience",
        ]

        wb = Workbook()

        ws = wb.active

        ws.title = "Employee Upload"

        ws.append(headers)

        master = wb.create_sheet("Master")

        master_data = {
            "CountryCodes": country_codes,
            "EmployeeTypes": self.EMPLOYEE_TYPES,
            "Currencies": self.CURRENCIES,
            "Managers": manager_values,
            "EmploymentTypes": self.EMPLOYEE_TYPES,
            "WorkModes": self.WORK_MODES,
            "EmploymentStatuses": self.EMPLOYMENT_STATUSES,
            "BloodGroups": self.BLOOD_GROUPS,
            "Genders": self.GENDERS,
            "MaritalStatuses": self.MARITAL_STATUSES,
        }

        col = 1

        for range_name, values in master_data.items():

            master.cell(row=1, column=col, value=range_name)

            for i, value in enumerate(values, start=2):

                master.cell(row=i, column=col, value=value)

            if values:

                col_letter = get_column_letter(col)

                wb.defined_names.add(
                    DefinedName(
                        range_name,
                        attr_text=(
                            f"Master!"
                            f"${col_letter}$2:"
                            f"${col_letter}${len(values)+1}"
                        ),
                    )
                )

            col += 1

        dept_map = {}

        for designation_uuid, designation_name, department_uuid in designations:

            dept_map.setdefault(department_uuid, []).append(designation_name)

        for department_uuid, department_name in departments:

            safe_name = department_name.replace(" ", "_")

            master.cell(row=1, column=col, value=safe_name)

            des_list = dept_map.get(department_uuid, [])

            for i, des in enumerate(des_list, start=2):

                master.cell(row=i, column=col, value=des)

            if des_list:

                col_letter = get_column_letter(col)

                wb.defined_names.add(
                    DefinedName(
                        safe_name,
                        attr_text=(
                            f"Master!"
                            f"${col_letter}$2:"
                            f"${col_letter}${len(des_list)+1}"
                        ),
                    )
                )

            col += 1

        dept_names = [dept[1] for dept in departments]

        master.cell(row=1, column=col, value="Departments")

        for i, dept_name in enumerate(dept_names, start=2):

            master.cell(row=i, column=col, value=dept_name)

        if dept_names:

            col_letter = get_column_letter(col)

            wb.defined_names.add(
                DefinedName(
                    "Departments",
                    attr_text=(
                        f"Master!"
                        f"${col_letter}$2:"
                        f"${col_letter}${len(dept_names)+1}"
                    ),
                )
            )

        header_to_col = {header: i + 1 for i, header in enumerate(headers)}

        def add_dropdown(header, range_name):

            col_idx = header_to_col[header]

            col_letter = get_column_letter(col_idx)

            dv = DataValidation(
                type="list", formula1=f"={range_name}", allow_blank=True
            )

            ws.add_data_validation(dv)

            dv.add(f"{col_letter}2:{col_letter}500")

        add_dropdown("country_code", "CountryCodes")
        add_dropdown("employee_type", "EmployeeTypes")
        add_dropdown("currency", "Currencies")
        add_dropdown("reporting_manager_uuid", "Managers")
        add_dropdown("employment_type", "EmploymentTypes")
        add_dropdown("work_mode", "WorkModes")
        add_dropdown("employment_status", "EmploymentStatuses")
        add_dropdown("blood_group", "BloodGroups")
        add_dropdown("gender", "Genders")
        add_dropdown("marital_status", "MaritalStatuses")
        add_dropdown("department", "Departments")

        dept_col = get_column_letter(header_to_col["department"])

        des_col = get_column_letter(header_to_col["designation"])

        for row in range(2, 501):

            dv = DataValidation(
                type="list",
                formula1=(f"=INDIRECT(" f'SUBSTITUTE(${dept_col}{row}," ","_"))'),
                allow_blank=True,
            )

            ws.add_data_validation(dv)

            dv.add(f"{des_col}{row}")

        master.sheet_state = "hidden"

        stream = BytesIO()

        wb.save(stream)

        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": "attachment; " "filename=employee_template.xlsx"
            },
        )
